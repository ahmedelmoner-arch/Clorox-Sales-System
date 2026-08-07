from pathlib import Path
from openpyxl import load_workbook
path = Path(r'c:\Users\elmon\OneDrive\سطح المكتب\Sales Usher (3).xlsx')
wb = load_workbook(path, data_only=True, read_only=True)
for sheet in ['Targets', 'Products']:
    ws = wb[sheet]
    print(f'--- {sheet} ---')
    for row in ws.iter_rows(min_row=1, max_row=8, values_only=True):
        print(row)
    print()