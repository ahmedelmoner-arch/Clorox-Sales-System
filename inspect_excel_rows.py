from pathlib import Path
from openpyxl import load_workbook
path = Path(r'c:\Users\elmon\OneDrive\سطح المكتب\Sales Usher (3).xlsx')
wb = load_workbook(path, data_only=True)
ws = wb['Targets']
print('max_row', ws.max_row)
print('max_column', ws.max_column)
print('header', next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
for row_num in range(1, 6):
    row = next(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True))
    print(row_num, row)
