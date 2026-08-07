from pathlib import Path
from openpyxl import load_workbook
path = Path(r'c:\Users\elmon\OneDrive\سطح المكتب\Sales Usher (3).xlsx')
wb = load_workbook(path, data_only=True)
ws = wb['Targets']
headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
product_columns = [i for i, header in enumerate(headers) if header not in {'UUID', 'Aug', 'Date', 'DelegateID', 'DelegateName', 'BranchID', 'BranchName', 'TargetConsumer'} and header]
print('product columns count', len(product_columns))
found = 0
for row_num, row in enumerate(ws.iter_rows(min_row=2, max_row=5000, values_only=True), start=2):
    values = [row[i] for i in product_columns]
    if any(v not in (None, '') for v in values):
        print('row', row_num, row[0], row[1], row[2], 'targets', [(headers[i], row[i]) for i in product_columns if row[i] not in (None, '')])
        found += 1
        if found >= 20:
            break
if not found:
    print('no product targets found in first 5000 rows')
