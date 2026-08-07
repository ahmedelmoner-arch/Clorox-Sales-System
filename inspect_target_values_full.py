from pathlib import Path
from openpyxl import load_workbook
path = Path(r'c:\Users\elmon\OneDrive\سطح المكتب\Sales Usher (3).xlsx')
wb = load_workbook(path, data_only=True, read_only=True)
ws = wb['Targets']
headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
known = {'UUID', 'Aug', 'Date', 'DelegateID', 'DelegateName', 'BranchID', 'BranchName', 'TargetConsumer'}
product_columns = [i for i, header in enumerate(headers) if header and header not in known]
print('header count', len(headers), 'product columns', len(product_columns))
for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    if row_num % 10000 == 0:
        print('scanned', row_num)
    if row is None:
        continue
    values = [row[i] if i < len(row) else None for i in product_columns]
    if any(v not in (None, '') for v in values):
        print('found row', row_num, row[0:8], [(headers[i], row[i]) for i in product_columns if row[i] not in (None, '')])
        break
else:
    print('none found')
