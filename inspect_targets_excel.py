from pathlib import Path
from openpyxl import load_workbook

path = Path(r'c:\Users\elmon\OneDrive\سطح المكتب\Sales Usher (3).xlsx')
wb = load_workbook(path, data_only=True, read_only=True)
print('sheets:', wb.sheetnames)
print('--- Targets ---')
ws = wb['Targets']
for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
    print(row)
